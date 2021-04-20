import openpyxl


def getrowcount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)
def getcolcount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)
def readdata(file, sheetName, rowno, colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowno,column=colno).value
def writedata(file,sheetName,rowno,colno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rowno,column=colno).value =data
    workbook.save(file)